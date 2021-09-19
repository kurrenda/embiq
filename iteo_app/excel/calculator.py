import openpyxl
from openpyxl.cell import cell as openpyxl_cell


class ExcelCalc():
    def __init__(self, filename, file_bytes, columns):
        self.filename = filename
        self.workbook = openpyxl.load_workbook(file_bytes)
        self.columns = columns

    @property
    def sheets(self):
        return self.workbook.sheetnames

    @staticmethod
    def parse_columns(value):
        input_columns = [x.strip() for x in value.split(',')]
        return input_columns

    @staticmethod
    def get_column_merge_width(sheet, cell):
        for merged_cell in sheet.merged_cells.ranges:
            if (cell.coordinate in merged_cell):
                merge_width = ColumnMergeWidth(
                    merged_cell.min_col,
                    merged_cell.max_col
                )
                return merge_width
        return False

    @staticmethod
    def get_columns_range(column_header):
        columns_to_check = []
        if column_header.merge_width:
            width = column_header.merge_width
            if width.col_min_width and width.col_max_width:
                for col in range(width.col_min_width, width.col_max_width):
                    columns_to_check.append(col)
        else:
            if column_header.col_coordinate:
                columns_to_check.append(column_header.col_coordinate)

        start_column = min(columns_to_check)
        end_column = max(columns_to_check)

        return start_column, end_column

    @staticmethod
    def _get_header_info(worksheet, input_column_names):
        row_count = worksheet.sheet.max_row

        for input_col_name in input_column_names:
            for worksheet_row in worksheet.sheet.iter_rows(1, row_count):
                for worksheet_col in worksheet_row:
                    col_val = worksheet_col.value
                    if col_val and isinstance(col_val, str):
                        if input_col_name == col_val.strip():
                            header_column_info = HeaderColumn(
                                worksheet_col.column,
                                worksheet_col.row
                            )
                            col_merge_width = ExcelCalc.get_column_merge_width(
                                worksheet.sheet,
                                worksheet_col
                            )

                            if col_merge_width:
                                header_column_info.merge_width = col_merge_width

                            column = Column(input_col_name)
                            column.header = header_column_info
                            worksheet.columns.append(column)

    @staticmethod
    def _prepare_response_data(worksheets, filename):

        response = {
            "file": filename,
            "summary": []
        }

        for worksheet in worksheets:
            for column in worksheet.columns:
                for stat in column.stats:
                    if len(column.stats) > 1:
                        column_name = f'{column.name} - {stat.coords}'
                    else:
                        column_name = column.name

                    stat = {
                            "column": column_name,
                            "sum": stat.sum,
                            "avg": stat.avg
                    }

                    if stat['sum'] and stat['avg']:
                        response['summary'].append(stat)

        return response

    def compute_data(self):
        worksheets = []

        for sheet_name in self.sheets:

            input_column_names = ExcelCalc.parse_columns(self.columns)

            worksheet = Worksheet(
                sheet_name,
                self.workbook[sheet_name]
            )

            self._get_header_info(worksheet, input_column_names)
            self._get_stats(worksheet)

            worksheets.append(worksheet)

        response = self._prepare_response_data(worksheets, self.filename)
        return response

    def _get_stats(self, worksheet):
        for column in worksheet.columns:
            if column.header:
                col_start_range, col_end_range = self.get_columns_range(column.header)
                col_iterator = worksheet.sheet.iter_cols(
                    min_col=col_start_range,
                    max_col=col_end_range,
                    min_row=column.header.row_coordinate)
                for col in col_iterator:
                    cell_sum = None
                    cell_count = None
                    first_cell_coords = None
                    for cell in col:
                        if cell.data_type == openpyxl_cell.TYPE_NUMERIC:
                            if cell.value is not None:
                                if cell_sum is None:
                                    first_cell_coords = cell.coordinate
                                    cell_sum = 0
                                if cell_count is None:
                                    cell_count = 0

                                cell_sum += cell.value
                                cell_count += 1
                    stats = Stats(
                        coords=first_cell_coords,
                        sum=cell_sum,
                        quantity=cell_count
                    )
                    column.stats.append(stats)


class Worksheet():
    def __init__(self, name, sheet):
        self.name = name
        self.sheet = sheet
        self.columns = []


class Column():
    def __init__(self, name):
        self.name = name
        self.header = None
        self.stats = []


class Stats():
    def __init__(self, coords, sum, quantity):
        self.coords = coords
        self._sum = sum
        self.quantity = quantity
        self._avg = None

    @property
    def avg(self):
        if self.sum and self.quantity:
            if self.quantity > 0:
                return round(self.sum/self.quantity, 2)
        return None

    @property
    def sum(self):
        if self._sum:
            return round(self._sum, 2)
        return None


class HeaderColumn():
    def __init__(self, col_coordinate, row_coordinate):
        self._col_coordinate = col_coordinate
        self._row_coordinate = row_coordinate
        self._merge_width = None

    @property
    def col_coordinate(self):
        return self._col_coordinate

    @property
    def row_coordinate(self):
        return self._row_coordinate

    @property
    def merge_width(self):
        return self._merge_width

    @merge_width.setter
    def merge_width(self, value):
        self._merge_width = value


class ColumnMergeWidth():
    def __init__(self, col_min_width, col_max_width):
        self._col_min_width = col_min_width
        self._col_max_width = col_max_width

    @property
    def col_min_width(self):
        return self._col_min_width

    @property
    def col_max_width(self):
        return self._col_max_width
